import streamlit as st
import pandas as pd
import openpyxl
import datetime
import io
import re

# Page Configuration & Styling
st.set_page_config(
    page_title="Sales Order Automation Hub", 
    page_icon="🚀", 
    layout="centered"
)

st.markdown("""
    <style>
        #GithubIcon { visibility: hidden; }
        .stButton>button {
            width: 100%;
            background-color: #10b981 !important;
            color: #ffffff !important;
            font-size: 16px;
            font-weight: 700;
            padding: 14px;
            border-radius: 8px;
            border: none;
        }
        .stButton>button:hover {
            background-color: #059669 !important;
        }
    </style>
""", unsafe_allow_html=True)

st.title("📊 Sales Order Automation Hub")
st.markdown("Upload multiple **Inbound Demand Files** to process orders in batch (Template is loaded locally).")
st.markdown("---")

# Session State for persistent download buttons
if 'processed_files' not in st.session_state:
    st.session_state.processed_files = []

# File Upload Section
uploaded_inputs = st.file_uploader("Upload Multiple Demand Excel Files", type=["xlsx", "xls"], accept_multiple_files=True, key="inputs")

st.markdown("<br>", unsafe_allow_html=True)

if st.button("🚀 Process Batch Orders", type="primary"):
    if uploaded_inputs:
        st.session_state.processed_files = []
        with st.spinner("⚡ Reading template and processing files... Please wait."):
            try:
                # Private repository ke liye local file read karne ka tareeqa
                try:
                    with open("Output.xlsx", "rb") as f:
                        template_bytes = f.read()
                except FileNotFoundError:
                    st.error("❌ 'Output.xlsx' template file repository mein nahi mili. Kripya template file ko GitHub repo ke main folder mein upload karein.")
                    st.stop()
                
                total_processed = 0
                total_orders_created = 0
                
                today_date = datetime.date.today().strftime("%Y-%m-%d")
                timestamp = datetime.datetime.now().strftime("%H%M%S")

                for uploaded_file in uploaded_inputs:
                    short_filename = uploaded_file.name
                    if short_filename.lower() == "output.xlsx":
                        continue

                    # Read input file via pandas
                    file_bytes = uploaded_file.getvalue()
                    df_input = pd.read_excel(io.BytesIO(file_bytes), header=None)

                    # 1. Find FG Row & Col
                    fg_row, fg_col = -1, -1
                    for r in range(df_input.shape[0]):
                        for c in range(df_input.shape[1]):
                            val = str(df_input.iloc[r, c]).strip().upper()
                            if val.startswith("FG"):
                                fg_row, fg_col = r, c
                                break
                        if fg_row != -1:
                            break

                    if fg_row == -1:
                        continue

                    # 2. Total/Sum Column Detection
                    total_col = df_input.shape[1]
                    for cSearch in range(fg_col, df_input.shape[1]):
                        h_val = str(df_input.iloc[fg_row, cSearch] if fg_row >= 0 else "").strip().upper()
                        h_prev = str(df_input.iloc[fg_row - 1, cSearch] if fg_row > 0 else "").strip().upper()
                        
                        is_total = "TOTAL" in h_val or "SUM" in h_val or "TOTAL" in h_prev or "SUM" in h_prev
                        if not is_total:
                            try:
                                cell_formula = str(df_input.iloc[fg_row + 1, cSearch]).upper()
                                if "SUM" in cell_formula:
                                    is_total = True
                            except:
                                pass
                        
                        if is_total:
                            total_col = cSearch
                            break

                    # 3. Route Number Finding Logic
                    route_num = "22"
                    ignore_list = ["RT", "DR", "RT DR", "ROUTE", "SALES PERSON", "CONTACT NO:", "MATERIAL CODE"]
                    
                    for r in range(min(fg_row, 5)):
                        for c in range(min(total_col, 30)):
                            cell_val = str(df_input.iloc[r, c]).strip()
                            upper_val = cell_val.upper()
                            
                            if upper_val in ignore_list:
                                continue
                                
                            is_product_code = any(upper_val.startswith(p) for p in ["PC", "MS", "M", "GM", "DP", "SKU", "FG"])
                            if is_product_code:
                                continue
                            
                            if cell_val != "" and 1 <= len(cell_val) <= 5:
                                if any(char.isdigit() for char in cell_val):
                                    route_num = cell_val
                                    break
                        if route_num != "22":
                            break

                    safe_route_num = "".join(c if c.isalnum() or c in ('-', '_') else "-" for c in str(route_num))

                    # 4. Smart Agency Detection (With Serial/Sequence Number Filtering)
                    agency_col = -1
                    for cSearch in range(fg_col - 1, -1, -1):
                        valid_agency_count = 0
                        
                        header_val = str(df_input.iloc[fg_row, cSearch] if fg_row >= 0 else "").strip().upper()
                        if any(s in header_val for s in ["S.NO", "SR", "NO.", "INDEX", "SEQ"]):
                            continue 

                        extracted_numbers = []
                        for rCheck in range(fg_row + 1, df_input.shape[0]):
                            v = df_input.iloc[rCheck, cSearch]
                            if pd.notna(v) and str(v).strip() != "":
                                clean_v = str(v).replace('.0', '').strip()
                                if clean_v.isdigit() and 1 <= len(clean_v) <= 5:
                                    extracted_numbers.append(int(clean_v))
                                    valid_agency_count += 1
                        
                        if len(extracted_numbers) > 2:
                            is_sequential = all(extracted_numbers[i] < extracted_numbers[i+1] for i in range(len(extracted_numbers)-1))
                            first_num = extracted_numbers[0]
                            if is_sequential and (first_num == 1 or first_num == 0):
                                continue 

                        if valid_agency_count > 0:
                            agency_col = cSearch
                            break

                    if agency_col == -1 and fg_col > 0:
                        agency_col = fg_col - 1

                    # 4.1 Strict DR Code Column Detection (Starting right below FG Row, format: DR + Numbers)
                    dr_code_col = -1
                    for cSearch in range(fg_col - 1, -1, -1):
                        sample_val = str(df_input.iloc[fg_row + 1, cSearch] if fg_row + 1 < df_input.shape[0] else "").strip().upper()
                        
                        if re.match(r'^DR\d+', sample_val):
                            dr_code_col = cSearch
                            break
                        
                        matched_count = 0
                        for offset in range(1, min(4, df_input.shape[0] - fg_row)):
                            v = str(df_input.iloc[fg_row + offset, cSearch]).strip().upper()
                            if re.match(r'^DR\d+', v):
                                matched_count += 1
                        
                        if matched_count > 0:
                            dr_code_col = cSearch
                            break

                    # 5. Valid FG Columns
                    valid_cols = []
                    for c in range(fg_col, total_col):
                        fg_code = str(df_input.iloc[fg_row, c] if fg_row >= 0 else "").strip()
                        valid_cols.append((c, fg_code))

                    # 6. Load Template for Valid & Missing DR Orders separately
                    wb_valid = openpyxl.load_workbook(io.BytesIO(template_bytes))
                    ws_valid = wb_valid["Order Data"] if "Order Data" in wb_valid.sheetnames else wb_valid.active

                    wb_missing = openpyxl.load_workbook(io.BytesIO(template_bytes))
                    ws_missing = wb_missing["Order Data"] if "Order Data" in wb_missing.sheetnames else wb_missing.active

                    valid_row = 6
                    missing_row = 6
                    valid_order_num = 1
                    missing_order_num = 1
                    
                    agency_counts_valid = {}
                    agency_counts_missing = {}
                    
                    valid_items_created = 0
                    missing_items_created = 0

                    for r in range(fg_row + 1, df_input.shape[0]):
                        agency = df_input.iloc[r, agency_col] if agency_col >= 0 else None
                        if pd.notna(agency) and str(agency).strip() != "":
                            agency_str = str(agency).replace('.0','').strip()
                            if agency_str.isdigit() and 1 <= len(agency_str) <= 5:
                                agency_val = int(agency_str)
                                
                                # Check if DR Code exists for this row
                                has_dr_code = False
                                clean_dr = ""
                                if dr_code_col >= 0:
                                    raw_dr = df_input.iloc[r, dr_code_col]
                                    if pd.notna(raw_dr) and str(raw_dr).strip() != "":
                                        clean_dr = str(raw_dr).replace('.0', '').strip()
                                        if clean_dr.upper() != "NAN" and clean_dr != "":
                                            has_dr_code = True

                                # Route based on DR Code presence
                                if has_dr_code:
                                    agency_counts_valid[agency_val] = agency_counts_valid.get(agency_val, 0) + 1
                                    current_seq = agency_counts_valid[agency_val]
                                    ref_number = f"RT-{route_num}-{agency_val}-{today_date}" if current_seq == 1 else f"RT-{route_num}-{agency_val}-{today_date}-{current_seq}"
                                    
                                    target_ws = ws_valid
                                    current_r = valid_row
                                    order_num = valid_order_num
                                    dr_to_use = clean_dr
                                else:
                                    # Missing DR Code (New Customer Case)
                                    agency_counts_missing[agency_val] = agency_counts_missing.get(agency_val, 0) + 1
                                    current_seq = agency_counts_missing[agency_val]
                                    ref_number = f"RT-{route_num}-{agency_val}-{today_date}-NEW" if current_seq == 1 else f"RT-{route_num}-{agency_val}-{today_date}-NEW-{current_seq}"
                                    
                                    target_ws = ws_missing
                                    current_r = missing_row
                                    order_num = missing_order_num
                                    dr_to_use = f"NEW_CUST_{agency_val}"

                                item_id = 10
                                row_has_items = False
                                
                                for c, fg_code in valid_cols:
                                    sku_qty = df_input.iloc[r, c]
                                    if pd.notna(sku_qty) and str(sku_qty).strip() != "":
                                        try:
                                            qty_val = float(sku_qty)
                                            if qty_val > 0:
                                                row_has_items = True
                                                current_fg = fg_code if (fg_code != "" and fg_code.lower() != "nan" and fg_code.upper().startswith("FG")) else "FG500014"
                                                
                                                target_ws.cell(row=current_r, column=2, value=order_num)
                                                target_ws.cell(row=current_r, column=3, value="OR")
                                                target_ws.cell(row=current_r, column=4, value="SO20")
                                                target_ws.cell(row=current_r, column=5, value=10)
                                                target_ws.cell(row=current_r, column=6, value=20)
                                                target_ws.cell(row=current_r, column=7, value=dr_to_use)
                                                target_ws.cell(row=current_r, column=8, value=dr_to_use)
                                                target_ws.cell(row=current_r, column=9, value=ref_number)
                                                target_ws.cell(row=current_r, column=10, value=today_date)
                                                target_ws.cell(row=current_r, column=11, value=today_date)
                                                target_ws.cell(row=current_r, column=15, value=item_id)
                                                target_ws.cell(row=current_r, column=16, value=current_fg)
                                                target_ws.cell(row=current_r, column=19, value=qty_val)
                                                target_ws.cell(row=current_r, column=20, value="Bag")
                                                target_ws.cell(row=current_r, column=22, value=2100)
                                                target_ws.cell(row=current_r, column=26, value=str(route_num))
                                                target_ws.cell(row=current_r, column=27, value=agency_val)
                                                
                                                item_id += 10
                                                current_r += 1
                                        except ValueError:
                                            pass
                                
                                if row_has_items:
                                    if has_dr_code:
                                        valid_row = current_r
                                        valid_order_num += 1
                                        valid_items_created += 1
                                    else:
                                        missing_row = current_r
                                        missing_order_num += 1
                                        missing_items_created += 1

                    if valid_items_created > 0:
                        buf_valid = io.BytesIO()
                        wb_valid.save(buf_valid)
                        buf_valid.seek(0)
                        st.session_state.processed_files.append({
                            "name": short_filename + " (Valid DR)",
                            "data": buf_valid.getvalue(),
                            "filename": safe_route_num + "_" + today_date + "_" + timestamp + "_Valid.xlsx",
                            "orders": valid_items_created
                        })

                    if missing_items_created > 0:
                        buf_missing = io.BytesIO()
                        wb_missing.save(buf_missing)
                        buf_missing.seek(0)
                        st.session_state.processed_files.append({
                            "name": short_filename + " (Missing DR / New Customer)",
                            "data": buf_missing.getvalue(),
                            "filename": safe_route_num + "_" + today_date + "_" + timestamp + "_Missing_DR.xlsx",
                            "orders": missing_items_created
                        })

                    total_processed += 1

                st.success("✅ Batch Processing Complete!")

            except Exception as e:
                st.error("❌ Error aagaya: " + str(e))
    else:
        st.warning("⚠️ Kripya pehle demand files upload karein!")

if st.session_state.processed_files:
    st.markdown("---")
    for item in st.session_state.processed_files:
        st.success("✅ Processed: " + item['name'] + " -> Orders created: " + str(item['orders']))
        if st.download_button(
            label="📥 Download " + item['name'],
            data=item['data'],
            file_name=item['filename'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=item['filename']
        ):
            st.toast(f"🎉 '{item['filename']}' successfully download ho gaya hai!", icon="📥")
    
    st.markdown("---")
    st.info("📊 **Batch Summary:** Total Output Files Generated: " + str(len(st.session_state.processed_files)))


